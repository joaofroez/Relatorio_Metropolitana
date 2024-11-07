import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog, messagebox

def ajustar_coluna_local(df):
    if 'Local' in df.columns:
        df[['Parte1', 'Parte2', 'Parte3','Parte4']] = df['Local'].str.split('/', expand=True, n=3)
   
        df = df.drop(columns=['Parte1', 'Parte2', 'Local','Parte4'], inplace=True)
    return df

def renomear_col(df):
    df.rename(columns={'Parte3':'Local'}, inplace=True)

def abreviar_data(df):
    df['Criada em'] = df['Criada em'].dt.date

def add_gerencias(df, df2):
    df = pd.merge(df, df2, on='Local', how='left')
    return df

def filtrar_metropolitana(df):
   df_metropolitana = df[df['GERÊNCIA'].isna()]
   return df_metropolitana 

def ajustar_posicao_col(df):
    drm = df.columns[-1]
    df_metropolitana_ordenada = list(df.columns[:-1])
    df_metropolitana_ordenada.insert(4, drm)
    df = df[df_metropolitana_ordenada]
    return df
def alterar_erros(df):
    df['GERÊNCIA'] = df['GERÊNCIA'].fillna('GRIO')

def add_nova_linha(df):
    total_linhas = len(df)
    cabecalho_linha = pd.DataFrame([df.columns], columns=df.columns)
    linha_vazia = pd.DataFrame([[''] * len(df.columns)], columns=df.columns)   
    nova_linha = pd.DataFrame([['Solicitações Comerciais Área Metropolitana', '', '', '', 'Total:', total_linhas, '', '', '', '', '']],
                          columns=df.columns)
    df = pd.concat([nova_linha, linha_vazia, cabecalho_linha, df_metropolitana], ignore_index=True)
    df.columns = [None] * df.shape[1]
    return df

def num_format_texto(df):
    df['Nº Ordem de Serviço Interna'] = df['Nº Ordem de Serviço Interna'].apply(lambda x: f"'{x}")

def criar_tabela(df):
    num_linhas = len(df) + 1
    num_colunas = len(df.columns)
    ultimo_nome_coluna = chr(64 + num_colunas) 
    ref = f"A4:{ultimo_nome_coluna}{num_linhas}"
    tabela = Table(displayName="Tabela1", ref=ref)
    estilo = TableStyleInfo(
    name="TableStyleMedium9",  
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
    )
    tabela.tableStyleInfo = estilo
    ws.add_table(tabela)

def add_img(caminho):
    img = Image(caminho)
    ws.add_image(img, 'A1')

def tirar_bordas_e_grades():    
    intervalo = 'A1:K1'
    bordas = Border(left=Side(style='none'), 
                     right=Side(style='none'), 
                     top=Side(style='none'), 
                     bottom=Side(style='none'))
    for linha in ws[intervalo]:
        for cell in linha:
            cell.border = bordas

def aumentar_dimensao_col_linha():
   for linha in ws.iter_rows(min_row=1, max_row=2):  
    ws.row_dimensions[linha[0].row].height = 60  
    
    for col in ws.columns:
        largura_max = 0
        col = [cell for cell in col]  
        for cell in col:
            try:
                if len(str(cell.value)) > largura_max:
                    largura_max = len(str(cell.value))  
            except:
                pass
        largura_ajus = (largura_max + 3)  
        ws.column_dimensions[col[0].column_letter].width = largura_ajus
def alterar_format_cel(*cells):
    estilo_fonte = Font(size=24, color='FFFFFF')
    estilo_fundo = PatternFill(start_color='5E90D2', end_color='5E90D2', fill_type='solid')
    
    for cell in cells:
        cell.font = estilo_fonte
        cell.fill = estilo_fundo

def juntar_cells(cell1,cell2):
    conteudo_junto = f"{ws[cell1].value} {ws[cell2].value}"
    ws[cell1] = conteudo_junto
    ws[cell2].value = None
def alinhar_e_centralizar():
    centralizado = Alignment(horizontal='center', vertical='center')
    max_linhas = ws.max_row
    max_col = ws.max_column
    for linha in ws.iter_rows(min_row=1, max_row=max_linhas, min_col=1, max_col=max_col):
        for cell in linha:
            cell.alignment = centralizado


def selecionar_arquivo():
    global arquivo_path
    arquivo_path = filedialog.askopenfilename(
        title="Selecione um arquivo .xlsx",
        filetypes=[("Arquivo Excel", "*.xlsx")]
    )
    if arquivo_path:
        messagebox.showinfo("Arquivo Selecionado", f"Arquivo selecionado:\n{arquivo_path}")
        janela.destroy()
        processar_arquivo()
    else:
        messagebox.showwarning("Nenhum arquivo", "Nenhum arquivo foi selecionado.")

def processar_arquivo():
    global df_metropolitana, wb, ws
#----------------------------CARREGAMENTO DOS ARQUVOS------------------------------------------------#
    CAMINHO_CODIGO = os.path.abspath(__file__)
    PASTA_CODIGO = os.path.dirname(CAMINHO_CODIGO)
    PASTA_PROGRAMA = os.path.dirname(PASTA_CODIGO)
    
    caminho_img = os.path.join(PASTA_PROGRAMA, 'img', 'cedae_img.jpg')
    caminho_arquivo = arquivo_path
    caminho_arquivo_procx = os.path.join(PASTA_PROGRAMA, 'base_sheets', 'Gerencias_fic.xlsx')

 
    df = pd.read_excel(caminho_arquivo, engine='openpyxl')
    df_gerencias = pd.read_excel(caminho_arquivo_procx, engine='openpyxl')

#----------------------------MANIPULAÇÃO DOS DADOS .XLSX--------------------------------------------------#
    ajustar_coluna_local(df)
    renomear_col(df)
    abreviar_data(df)
    df = add_gerencias(df, df_gerencias)
    df_metropolitana = filtrar_metropolitana(df)
    df_metropolitana = ajustar_posicao_col(df_metropolitana)
    alterar_erros(df_metropolitana)
    num_format_texto(df_metropolitana)
    df_metropolitana = add_nova_linha(df_metropolitana)
#----------------------------SALVAMENTO E CARREGAMENTO DOS DADOS PROCESSADOS-------------------------------#
    temp_sheet = os.path.join(PASTA_PROGRAMA, 'processed-sheets', 'temp_sheet.xlsx')
    df_metropolitana.to_excel(temp_sheet, index=False)

    wb = load_workbook(temp_sheet)
    ws = wb.active

#----------------------------FORMATAÇÃO DA PLANILHA--------------------------------------------------------#
    criar_tabela(df_metropolitana)
    add_img(caminho_img)
    tirar_bordas_e_grades()
    aumentar_dimensao_col_linha()
    alterar_format_cel(ws['A2'], ws['B2'], ws['C2'], ws['D2'], ws['E2'], ws['F2'])
    juntar_cells('E2', 'F2')
    ws.merge_cells('A2:C2')
    ws.merge_cells('E2:F2')
    alinhar_e_centralizar()
    ws.sheet_view.showGridLines = False

   
    salvar_arquivo_processado()
#------------------------SALVAMENTO DO ARQUIVO FORMATADO----------------------------#
def salvar_arquivo_processado():

    local_salvo = filedialog.asksaveasfilename(
        title="Salvar arquivo como",
        defaultextension=".xlsx",
        filetypes=[("Arquivo Excel", "*.xlsx")]
    )
    if local_salvo:
        wb.save(local_salvo)  
        messagebox.showinfo("Arquivo Salvo", f"O arquivo foi salvo em:\n{local_salvo}")
    else:
        messagebox.showwarning("Operação Cancelada", "O arquivo não foi salvo.")

#--------------------INTERFACE GRÁFICA----------------------------------------------#
janela = tk.Tk()
janela.title("Gerador de relatórios")
janela.geometry("500x200")

titulo = tk.Label(
    janela,
    text="Solicitações Comerciais Área Metropolitana",
    font=("Arial", 16, "bold"),
    fg="#333333"
)
titulo.pack(pady=(20, 10))

botao = tk.Button(
    janela, 
    text="Selecionar Arquivo .xlsx", 
    command=selecionar_arquivo,
    font=("Arial", 12, "bold"),
    bg="#4CAF50",
    fg="white",
    padx=20,
    pady=10
)
botao.pack(pady=50)

janela.mainloop()